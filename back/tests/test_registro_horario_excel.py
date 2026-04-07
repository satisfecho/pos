"""Spanish registro horario Excel export (reports)."""
from __future__ import annotations

import io
from datetime import datetime, timedelta, timezone

from openpyxl import load_workbook
from pg_client_mixin import PgClientTestCase

from app import models, security


def _bearer_headers(user: models.User) -> dict[str, str]:
    data = {
        "sub": user.email,
        "tenant_id": user.tenant_id,
        "provider_id": getattr(user, "provider_id", None),
        "token_version": user.token_version,
    }
    token = security.create_access_token(data, expires_delta=timedelta(minutes=30))
    return {"Authorization": f"Bearer {token}"}


class TestRegistroHorarioExcel(PgClientTestCase):
    def setUp(self) -> None:
        super().setUp()
        tenant = models.Tenant(name="RH Test", cif="B12345678", address="Calle 1", ccc="12/12345678/12")
        self.session.add(tenant)
        self.session.commit()
        self.session.refresh(tenant)
        self.tenant_id = tenant.id

        self.staff_a = models.User(
            email="rh-staff-a@test.local",
            hashed_password=security.get_password_hash("secret"),
            full_name="Staff Alpha",
            employee_number="E-1",
            tenant_id=self.tenant_id,
            role=models.UserRole.waiter,
        )
        self.admin = models.User(
            email="rh-admin@test.local",
            hashed_password=security.get_password_hash("secret"),
            full_name="Admin A",
            tenant_id=self.tenant_id,
            role=models.UserRole.admin,
        )
        self.session.add_all([self.staff_a, self.admin])
        self.session.commit()
        self.session.refresh(self.staff_a)
        self.session.refresh(self.admin)

        base = datetime(2026, 3, 15, 9, 0, tzinfo=timezone.utc)
        ws = models.WorkSession(
            tenant_id=self.tenant_id,
            user_id=self.staff_a.id,
            started_at=base,
            ended_at=base + timedelta(hours=4),
        )
        self.session.add(ws)
        self.session.commit()

    def test_registro_horario_excel_returns_200_and_signature_column(self) -> None:
        headers = _bearer_headers(self.admin)
        r = self.client.get(
            "/reports/attendance-registro-horario-excel",
            params={"year": 2026, "month": 3},
            headers=headers,
        )
        self.assertEqual(r.status_code, 200, r.text)
        self.assertIn("spreadsheetml", r.headers.get("content-type", ""))

        wb = load_workbook(io.BytesIO(r.content), read_only=True)
        self.assertEqual(len(wb.sheetnames), 1)
        sheet = wb[wb.sheetnames[0]]
        header_row = None
        for row in sheet.iter_rows(min_row=1, max_row=40, max_col=6):
            vals = [c.value for c in row]
            if vals[0] == "Día" and vals[-1] == "Firma del empleado":
                header_row = row[0].row
                break
        self.assertIsNotNone(header_row, "expected Spanish grid header row")
        # March has 31 days
        self.assertEqual(sheet.max_row >= header_row + 31, True)

    def test_registro_horario_invalid_staff_id_400(self) -> None:
        headers = _bearer_headers(self.admin)
        r = self.client.get(
            "/reports/attendance-registro-horario-excel",
            params={"year": 2026, "month": 3, "staff_ids": 999_999_999},
            headers=headers,
        )
        self.assertEqual(r.status_code, 400, r.text)
