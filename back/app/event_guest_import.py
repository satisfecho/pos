"""Event guest-list import: parse uploaded .xlsx/.csv, preview, confirm, and template.

This is the first server-side spreadsheet parser in the backend (product import parses on the
frontend). Mirrors the product bulk-import preview/confirm shape. Only `name` is mandatory.
"""

from __future__ import annotations

import csv
import io
import re
import unicodedata
import uuid
from datetime import datetime, timezone

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel, Field
from sqlmodel import Session, select

from . import models

MAX_GUEST_IMPORT_ROWS = 2000

# Downloadable template columns (header text shown to the organizer).
TEMPLATE_HEADERS = ["Nombre", "Telefono", "Email", "Acompanantes", "Mesa", "Notas"]
TEMPLATE_EXAMPLES = [
    ["Maria Pérez", "5512345678", "maria@correo.com", 2, "Mesa 1", "Vegetariana"],
    ["Juan López", "", "", 1, "", ""],
]

# Field <- header keyword aliases (checked in order; substring match on a normalized header).
_HEADER_RULES: list[tuple[str, tuple[str, ...]]] = [
    ("name", ("nombre", "name", "invitado", "guest", "fullname")),
    ("phone", ("telefono", "tel", "celular", "movil", "whats", "phone", "mobile", "cel")),
    ("email", ("email", "correo", "mail")),
    ("party_size", ("acompan", "acompaniantes", "+1", "plusone", "pax", "personas", "invitados")),
    ("table_label", ("mesa", "grupo", "table", "group")),
    ("notes", ("nota", "observ", "coment", "notes")),
]

_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _strip_accents(text: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFKD", text) if not unicodedata.combining(c)
    )


def normalize_name(name: str) -> str:
    """For duplicate matching: collapse whitespace, strip accents, casefold."""
    return _strip_accents(re.sub(r"\s+", " ", (name or "").strip())).casefold()


def _normalize_header(header: str) -> str:
    return re.sub(r"[\s_.\-]+", "", _strip_accents(str(header or "")).lower())


def _map_headers(headers: list[str]) -> dict[int, str]:
    """Map each column index to a guest field (first matching rule wins, no double-mapping)."""
    out: dict[int, str] = {}
    used: set[str] = set()
    for idx, raw in enumerate(headers):
        norm = _normalize_header(raw)
        if not norm:
            continue
        for field, keywords in _HEADER_RULES:
            if field in used:
                continue
            if any(k in norm for k in keywords):
                out[idx] = field
                used.add(field)
                break
    return out


def _coerce_party_size(value) -> int:
    if value is None:
        return 1
    try:
        n = int(float(str(value).strip()))
    except (ValueError, TypeError):
        return 1
    return n if n >= 1 else 1


def _clean(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


class GuestImportRow(BaseModel):
    row_index: int
    name: str = ""
    phone: str | None = None
    email: str | None = None
    party_size: int = 1
    table_label: str | None = None
    notes: str | None = None
    valid: bool = False
    errors: list[str] = Field(default_factory=list)  # blocking
    warnings: list[str] = Field(default_factory=list)  # non-blocking
    action: str = "skip"  # "create" | "skip"
    duplicate_of_row: int | None = None
    duplicate_existing: bool = False


class GuestImportSummary(BaseModel):
    total: int
    valid: int
    skipped: int
    duplicates: int
    create: int


class GuestImportPreview(BaseModel):
    items: list[GuestImportRow]
    summary: GuestImportSummary


def parse_spreadsheet(contents: bytes, filename: str) -> list[dict]:
    """Parse the first sheet of an .xlsx/.csv into a list of field dicts. Raises ValueError."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        rows = _parse_csv(contents)
    else:
        rows = _parse_xlsx(contents)
    if not rows:
        raise ValueError("empty_file")
    header = rows[0]
    mapping = _map_headers([str(c) if c is not None else "" for c in header])
    if "name" not in mapping.values():
        # No recognizable name column: assume the first column is the name.
        mapping[0] = "name"
    out: list[dict] = []
    for raw in rows[1:]:
        if raw is None:
            continue
        record: dict = {}
        any_value = False
        for idx, field in mapping.items():
            if idx >= len(raw):
                continue
            val = raw[idx]
            if field == "party_size":
                record[field] = _coerce_party_size(val)
            else:
                cleaned = _clean(val)
                record[field] = cleaned
                if cleaned:
                    any_value = True
        if not any_value and not record.get("name"):
            continue  # fully empty row
        out.append(record)
        if len(out) > MAX_GUEST_IMPORT_ROWS:
            raise ValueError("too_many_rows")
    return out


def _parse_xlsx(contents: bytes) -> list[tuple]:
    try:
        wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ValueError("invalid_xlsx") from exc
    try:
        ws = wb.worksheets[0]
        return [tuple(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _parse_csv(contents: bytes) -> list[tuple]:
    text: str | None = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = contents.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("invalid_encoding")
    reader = csv.reader(io.StringIO(text))
    return [tuple(row) for row in reader]


def build_guest_preview(
    session: Session,
    tenant_id: int,
    event_id: int,
    items: list[models.EventGuestImportItem],
) -> GuestImportPreview:
    existing_names = {
        normalize_name(g.name)
        for g in session.exec(
            select(models.EventGuest).where(
                models.EventGuest.tenant_id == tenant_id,
                models.EventGuest.event_id == event_id,
            )
        ).all()
        if g.name
    }

    seen_in_file: dict[str, int] = {}
    rows: list[GuestImportRow] = []
    summary = GuestImportSummary(total=len(items), valid=0, skipped=0, duplicates=0, create=0)

    for idx, item in enumerate(items):
        name = (item.name or "").strip()
        errors: list[str] = []
        warnings: list[str] = []
        if not name:
            errors.append("name_required")

        email = (item.email or "").strip() or None
        if email and not _EMAIL_RE.match(email):
            warnings.append("email_invalid")

        phone = (item.phone or "").strip() or None
        norm = normalize_name(name) if name else ""

        duplicate_of_row: int | None = None
        duplicate_existing = False
        if norm:
            if norm in existing_names:
                duplicate_existing = True
                warnings.append("duplicate_existing")
            elif norm in seen_in_file:
                duplicate_of_row = seen_in_file[norm]
                warnings.append("duplicate_in_file")
            else:
                seen_in_file[norm] = idx

        valid = not errors
        is_dup = duplicate_existing or duplicate_of_row is not None
        action = "create" if valid else "skip"

        rows.append(
            GuestImportRow(
                row_index=idx,
                name=name,
                phone=phone,
                email=email,
                party_size=_coerce_party_size(item.party_size),
                table_label=(item.table_label or "").strip() or None,
                notes=(item.notes or "").strip() or None,
                valid=valid,
                errors=errors,
                warnings=warnings,
                action=action,
                duplicate_of_row=duplicate_of_row,
                duplicate_existing=duplicate_existing,
            )
        )
        if valid:
            summary.valid += 1
            summary.create += 1
        else:
            summary.skipped += 1
        if is_dup:
            summary.duplicates += 1

    return GuestImportPreview(items=rows, summary=summary)


def confirm_guest_import(
    session: Session,
    tenant_id: int,
    event_id: int,
    items: list[models.EventGuestImportItem],
    skip_duplicates: bool = True,
) -> dict:
    """Persist valid rows. Returns {created, skipped}. Dedupe against existing + within batch."""
    existing_names = {
        normalize_name(g.name)
        for g in session.exec(
            select(models.EventGuest).where(
                models.EventGuest.tenant_id == tenant_id,
                models.EventGuest.event_id == event_id,
            )
        ).all()
        if g.name
    }
    created = 0
    skipped = 0
    now = datetime.now(timezone.utc)

    for item in items:
        name = (item.name or "").strip()
        if not name:
            skipped += 1
            continue
        norm = normalize_name(name)
        if skip_duplicates and norm in existing_names:
            skipped += 1
            continue
        guest = models.EventGuest(
            tenant_id=tenant_id,
            event_id=event_id,
            name=name,
            phone=(item.phone or "").strip() or None,
            email=(item.email or "").strip() or None,
            party_size=_coerce_party_size(item.party_size),
            table_label=(item.table_label or "").strip() or None,
            notes=(item.notes or "").strip() or None,
            status=models.InvitationStatus.pending,
            token=uuid.uuid4().hex,
            created_at=now,
            updated_at=now,
        )
        session.add(guest)
        existing_names.add(norm)
        created += 1

    session.commit()
    return {"created": created, "skipped": skipped}


def build_template_xlsx() -> bytes:
    """A one-sheet .xlsx with a highlighted required Nombre column + example rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Invitados"
    ws.append(TEMPLATE_HEADERS)

    header_fill = PatternFill(start_color="FFD9534F", end_color="FFD9534F", fill_type="solid")
    optional_fill = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
    bold_white = Font(bold=True, color="FFFFFFFF")
    bold_grey = Font(bold=True, color="FF555555")
    for col, header in enumerate(TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col)
        if header == "Nombre":
            cell.fill = header_fill
            cell.font = bold_white
        else:
            cell.fill = optional_fill
            cell.font = bold_grey

    for example in TEMPLATE_EXAMPLES:
        ws.append(example)

    widths = [24, 16, 26, 14, 12, 24]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
